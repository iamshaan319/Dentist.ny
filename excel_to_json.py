"""
excel_to_json.py
================
Converts the ny_dentists.xlsx output from Doctor.py into
the dentists.json format required by the website.

Usage:
    python excel_to_json.py ny_dentists.xlsx data/dentists.json

Requirements:
    pip install openpyxl
"""

import sys
import json
import re
import os
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def excel_to_json(input_xlsx: str, output_json: str):
    if not os.path.exists(input_xlsx):
        print(f"ERROR: File not found: {input_xlsx}")
        sys.exit(1)

    print(f"Reading {input_xlsx}…")
    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip().lower().replace(" ", "_").replace("/", "_")
               if h else f"col_{i}" for i, h in enumerate(headers)]

    dentists = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        raw = dict(zip(headers, row))

        name = str(raw.get("full_name", "") or "").strip()
        if not name:
            continue

        d = {
            "id": row_num - 1,
            "name": name,
            "first_name": str(raw.get("first_name", "") or "").strip(),
            "last_name":  str(raw.get("last_name",  "") or "").strip(),
            "degree":     str(raw.get("degree___credential", "") or raw.get("degree", "") or "").strip(),
            "specialty":  str(raw.get("specialty", "General Dentistry") or "General Dentistry").strip(),
            "practice_name": str(raw.get("practice___clinic_name", "") or raw.get("practice___clinic", "") or "").strip(),
            "address":    str(raw.get("address", "") or "").strip(),
            "city":       str(raw.get("city", "New York") or "New York").strip(),
            "state":      str(raw.get("state", "NY") or "NY").strip(),
            "zip":        str(raw.get("zip_code", "") or "").strip(),
            "borough":    _borough(str(raw.get("city", "") or ""), str(raw.get("zip_code", "") or "")),
            "phone":      str(raw.get("phone", "") or "").strip(),
            "website":    str(raw.get("website", "") or "").strip(),
            "email":      str(raw.get("email", "") or "").strip(),
            "accepting_new_patients": _bool(raw.get("accepting_new_patients")),
            "languages":  _list(raw.get("languages")),
            "gender":     str(raw.get("gender", "") or "").strip(),
            "npi":        str(raw.get("npi_number", "") or "").strip(),
            "rating":     _float(raw.get("rating")),
            "review_count": _int(raw.get("review_count")),
            "years_experience": _int(raw.get("years_of_experience")),
            "certifications": str(raw.get("certifications", "") or "").strip(),
            "office_hours": _parse_hours(raw.get("office_hours")),
            "insurance":  _list(raw.get("insurance_accepted")),
            "services":   [],
            "latitude":   _float(raw.get("latitude")),
            "longitude":  _float(raw.get("longitude")),
            "source":     str(raw.get("data_source", "") or "").strip(),
            "profile_url": str(raw.get("profile_url", "") or "").strip(),
            "image_placeholder": _initials(name),
        }

        # Clean up empty strings to None-equivalents
        for key in ["degree","practice_name","email","website","npi","certifications","source","profile_url"]:
            if d[key] == "None" or d[key] == "nan":
                d[key] = ""

        dentists.append(d)

    output = {
        "meta": {
            "total": len(dentists),
            "city": "New York",
            "state": "NY",
            "last_updated": str(date.today()),
            "sources": list({d["source"] for d in dentists if d["source"]}),
        },
        "dentists": dentists
    }

    # Create output directory if needed
    os.makedirs(os.path.dirname(output_json) if os.path.dirname(output_json) else ".", exist_ok=True)

    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"✅  Wrote {len(dentists)} dentists → {output_json}")
    print(f"    Boroughs: {sorted({d['borough'] for d in dentists})}")
    print(f"    Specialties: {len({d['specialty'] for d in dentists})}")


# ── Helpers ───────────────────────────────────────────────────────

def _bool(v):
    if v is None: return False
    if isinstance(v, bool): return v
    return str(v).strip().lower() in ("yes","true","1","accepting","y")

def _float(v):
    if v is None: return None
    try: return round(float(v), 6)
    except: return None

def _int(v):
    if v is None: return None
    try: return int(float(str(v).replace(",","")))
    except: return None

def _list(v):
    if v is None: return []
    s = str(v).strip()
    if not s or s.lower() in ("none","nan","n/a",""): return []
    return [x.strip() for x in re.split(r"[,;|]", s) if x.strip()]

def _parse_hours(v):
    if not v or str(v).lower() in ("none","nan",""): return {}
    parts = str(v).split("|")
    hours = {}
    for p in parts:
        p = p.strip()
        if ":" in p:
            day, time = p.split(":", 1)
            hours[day.strip()] = time.strip()
    return hours if hours else {}

def _borough(city: str, zip_code: str) -> str:
    city_lower = city.lower()
    if any(x in city_lower for x in ["manhattan","new york city","nyc"]):
        return "Manhattan"
    if "brooklyn" in city_lower:     return "Brooklyn"
    if "queens"   in city_lower:     return "Queens"
    if "bronx"    in city_lower:     return "Bronx"
    if "staten island" in city_lower: return "Staten Island"
    if "jackson heights" in city_lower or "flushing" in city_lower or "astoria" in city_lower:
        return "Queens"
    if "bay ridge" in city_lower or "greenpoint" in city_lower or "williamsburg" in city_lower:
        return "Brooklyn"
    # Try by ZIP code prefix
    z = str(zip_code)[:5]
    if z.startswith("100") or z.startswith("101") or z.startswith("102"):
        return "Manhattan"
    if z.startswith("112"): return "Brooklyn"
    if z.startswith("113") or z.startswith("114") or z.startswith("116"): return "Queens"
    if z.startswith("104"): return "Bronx"
    if z.startswith("103"): return "Staten Island"
    return city if city else "New York"

def _initials(name: str) -> str:
    words = [w for w in name.split() if w and not w.endswith(".")]
    if len(words) >= 3:
        return (words[1][0] + words[2][0]).upper()
    if len(words) == 2:
        return (words[0][0] + words[1][0]).upper()
    return name[:2].upper()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python excel_to_json.py ny_dentists.xlsx data/dentists.json")
        sys.exit(1)
    excel_to_json(sys.argv[1], sys.argv[2])
